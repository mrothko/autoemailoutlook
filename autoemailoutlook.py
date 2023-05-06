import openpyxl
import win32com.client as win32
import os

# 设置 Outlook 应用程序 ID 和机密
outlook_app_id = "your_outlook_app_id"
outlook_app_secret = "your_outlook_app_secret"

# 创建 Outlook 应用程序实例
outlook = win32.Dispatch("Outlook.Application")
mapi = outlook.GetNamespace("MAPI")

# 登录 Outlook 账户
account = mapi.LogonWithEWS("your_email_address", outlook_app_id, outlook_app_secret)

# 读取 Excel 文件中的收件人列表和邮件内容
excel_file_path = "email_list.xlsx"
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active
mail_subject = sheet["A1"].value
mail_body = sheet["A2"].value
mail_attachments = []

# 构建邮件附件列表
for row in sheet.iter_rows(min_row=3, max_col=1):
    attachment_path = row[0].value
    if attachment_path is not None and os.path.exists(attachment_path):
        mail_attachments.append(attachment_path)

# 获取 Outlook 联系人列表
contact_group_name = "My Contacts"
contact_group = None
for folder in account.Folders:
    if folder.Name == "Contacts":
        for contact_group_item in folder.Items:
            if contact_group_item.Name == contact_group_name:
                contact_group = contact_group_item
                break
        break

# 发送邮件
for row in sheet.iter_rows(min_row=3, max_col=2):
    recipient_email = row[0].value
    recipient_name = row[1].value
    if recipient_email is not None:
        recipient = outlook.CreateItem(0)
        recipient.To = recipient_email
        recipient.Subject = mail_subject
        recipient.Body = mail_body
        if len(mail_attachments) > 0:
            for attachment_path in mail_attachments:
                attachment = recipient.Attachments.Add(attachment_path)
                attachment.DisplayName = os.path.basename(attachment_path)
        if contact_group is not None:
            for contact in contact_group.Members:
                if contact.Email1Address == recipient_email:
                    recipient.CC = contact.Email2Address
                    break
        recipient.Send()

# 关闭 Outlook 应用程序实例
outlook.Quit()
